"""
app/services/gst_report_export.py
=====================================
Renders the dicts produced by app/services/gst_reports.py into an .xlsx
workbook (one sheet per GSTR-1 section) via openpyxl -- the only actual
use of that already-declared dependency (requirements.txt) in this repo.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_SECTION_COLUMNS: dict[str, list[str]] = {
    "b2b": ["gstin", "receiver_name", "invoice_number", "invoice_date", "invoice_value",
            "place_of_supply", "reverse_charge", "taxable_value", "tax_rate",
            "igst_amount", "cgst_amount", "sgst_amount"],
    "b2cl": ["gstin", "receiver_name", "invoice_number", "invoice_date", "invoice_value",
             "place_of_supply", "reverse_charge", "taxable_value", "tax_rate",
             "igst_amount", "cgst_amount", "sgst_amount"],
    "b2cs": ["place_of_supply", "tax_rate", "taxable_value", "igst_amount",
             "cgst_amount", "sgst_amount", "invoice_count"],
    "cdnr": ["gstin", "receiver_name", "note_number", "note_date", "note_type",
             "original_invoice_number", "place_of_supply", "note_value",
             "taxable_value", "igst_amount", "cgst_amount", "sgst_amount"],
    "cdnur": ["receiver_name", "note_number", "note_date", "note_type",
              "original_invoice_number", "place_of_supply", "note_value",
              "taxable_value", "igst_amount", "cgst_amount", "sgst_amount"],
    "hsn_summary": ["hsn_sac_code", "total_quantity", "taxable_value", "tax_rate",
                     "igst_amount", "cgst_amount", "sgst_amount", "total_value"],
}

_SHEET_TITLES = {
    "b2b": "B2B Invoices",
    "b2cl": "B2C Large",
    "b2cs": "B2C Small",
    "cdnr": "Credit-Debit Notes (Reg)",
    "cdnur": "Credit-Debit Notes (Unreg)",
    "hsn_summary": "HSN Summary",
}


def _write_sheet(ws: Worksheet, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col) for col in columns])


def render_gstr1_workbook(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for section, columns in _SECTION_COLUMNS.items():
        ws = wb.create_sheet(title=_SHEET_TITLES[section])
        _write_sheet(ws, columns, data.get(section, []))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_gstr3b_workbook(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR-3B Summary"
    for key, value in data.items():
        ws.append([key, value])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
